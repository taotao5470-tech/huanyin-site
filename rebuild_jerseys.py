#!/usr/bin/env python3
"""
Rebuild national team jersey data for World Cup site.
Consolidates 119 SPUs → ~53 SPUs (one per country).
Uses images from 球衣 (2).xlsx (main) and 球衣 (5).xlsx (model).
"""
import json
import os
import platform
import re
import shutil
import sys
import copy
from collections import defaultdict

# === CONFIG ===
# Detect if running on WSL (Linux) or Windows
if platform.system() == 'Linux':
    BASE = '/mnt/c/Users/Administrator/Desktop'
else:
    BASE = r'C:\Users\Administrator\Desktop'

SITE_DIR = os.path.join(BASE, '独立站备份', '世界杯策划版')
IMGS_DIR = os.path.join(SITE_DIR, 'images', 'mundial')
JERSEY2_DIR = os.path.join(IMGS_DIR, 'jersey2')
JERSEY5_DIR = os.path.join(IMGS_DIR, 'jersey5')

# === CN SKU name → Spanish team name ===
CN_TO_ES = {
    '德国': 'Alemania', '日本': 'Japón', '巴西': 'Brasil', '葡萄牙': 'Portugal',
    '墨西哥': 'México', '西班牙': 'España', '比利时': 'Bélgica', '阿根廷': 'Argentina',
    '挪威': 'Noruega', '法国': 'Francia', '加拿大': 'Canadá', '英格兰': 'Inglaterra',
    '乌拉圭': 'Uruguay', '苏格兰': 'Escocia', '意大利': 'Italia', '智利': 'Chile',
    '委内瑞拉': 'Venezuela', '荷兰': 'Países Bajos', '塞内加尔': 'Senegal',
    '威尔士': 'Gales', '瑞典': 'Suecia', '喀唛隆': 'Camerún', '卡塔尔': 'Catar',
    '沙特阿拉伯': 'Arabia Saudita', '沙特': 'Arabia Saudita', '瑞士': 'Suiza',
    '秘鲁': 'Perú', '克罗地亚': 'Croacia', '摩洛哥': 'Marruecos',
    '阿尔及利亚': 'Argelia', '爱尔兰': 'Irlanda', '北爱尔兰': 'Irlanda del Norte',
    '哥伦比亚': 'Colombia', '海地': 'Haití', '加纳': 'Ghana',
    '哥斯达黎加': 'Costa Rica', '科特迪瓦': 'Costa de Marfil', '巴拉圭': 'Paraguay',
    '美国': 'Estados Unidos', '尼日利亚': 'Nigeria', '牙买加': 'Jamaica',
    '中国': 'China', '中国队': 'China', '乌克兰': 'Ucrania', '佛得角': 'Cabo Verde',
    '匈牙利': 'Hungría', '南非': 'Sudáfrica', '土耳其': 'Turquía', '奥地利': 'Austria',
    '阿联酋长': 'Emiratos Árabes', '韩国': 'Corea', '马里': 'Malí',
    '葡萄牙黑豹': 'Portugal', '阿根廷赛前': 'Argentina',
    '刚果': 'Congo',
}

# === Region mapping ===
AMERICAS = {'México', 'Brasil', 'Argentina', 'Canadá', 'Uruguay', 'Chile',
            'Venezuela', 'Colombia', 'Haití', 'Estados Unidos', 'Jamaica',
            'Paraguay', 'Perú', 'Costa Rica', 'América', 'Congo'}
EUROPE = {'Alemania', 'Portugal', 'España', 'Bélgica', 'Inglaterra',
          'Escocia', 'Italia', 'Noruega', 'Francia', 'Países Bajos',
          'Gales', 'Suecia', 'Suiza', 'Irlanda', 'Irlanda del Norte',
          'Croacia', 'Polonia', 'Austria', 'Hungría', 'Ucrania', 'Turquía'}
AFRICA_ASIA = {'Japón', 'Senegal', 'Camerún', 'Catar', 'Arabia Saudita',
               'Marruecos', 'Argelia', 'Ghana', 'Costa de Marfil', 'Nigeria',
               'Cabo Verde', 'Sudáfrica', 'China', 'Corea', 'Malí',
               'Emiratos Árabes'}

def get_region(team_es):
    if team_es in AMERICAS: return 'americas'
    if team_es in EUROPE: return 'europe'
    if team_es in AFRICA_ASIA: return 'africa_asia'
    return 'europe'  # default

def extract_country_and_variant(sku_cn):
    """Parse SKU name like '2627德国主场' → (country_cn, variant_cn)"""
    # Remove season prefix
    name = re.sub(r'^(26|27|2526|2627|2026|2027)', '', sku_cn).strip()

    # Known variant suffixes
    variant_map = [
        ('二客', '二客场'), ('客场', '客场'), ('第二客场', '二客场'),
        ('主场', '主场'), ('特别版', '特别版'),
        ('球员版', '球员版'), ('赛前训练服', '赛前训练服'),
        ('训练服', '训练服'), ('守门员', '守门员'),
        ('白色', '特别版 - Blanco'), ('蓝色', '特别版 - Azul'),
        ('枣红色', '特别版 - Rojo'), ('黑色', '特别版 - Negro'),
        ('绿色', '特别版 - Verde'), ('粉色', '特别版 - Rosa'),
        ('款式1', '主场'), ('款式2', '客场'), ('款式3', '特别版 - 1'), ('款式4', '特别版 - 2'),
    ]

    variant_cn = '主场'  # default
    country_cn = name

    for suffix, variant in variant_map:
        if name.endswith(suffix):
            country_cn = name[:-len(suffix)]
            variant_cn = variant
            break

    # Clean up trailing season numbers on country
    country_cn = re.sub(r'^(26|27|2526|2627|2026|2027)', '', country_cn).strip()

    return country_cn, variant_cn

def variant_to_group(variant_cn):
    """Classify variant as 'home' or 'away'"""
    if '客场' in variant_cn:
        return 'away'
    return 'home'  # 主场, 特别版, 训练服 etc all go to home

def variant_label_es(variant_cn):
    if '客场' in variant_cn:
        return 'Visitante'
    if '特别版' in variant_cn:
        if 'Blanco' in variant_cn: return 'Ed. Especial - Blanco'
        if 'Azul' in variant_cn: return 'Ed. Especial - Azul'
        if 'Rojo' in variant_cn: return 'Ed. Especial - Rojo'
        if 'Negro' in variant_cn: return 'Ed. Especial - Negro'
        if 'Verde' in variant_cn: return 'Ed. Especial - Verde'
        if 'Rosa' in variant_cn: return 'Ed. Especial - Rosa'
        return 'Edición Especial'
    if '训练服' in variant_cn or '赛前' in variant_cn:
        return 'Entrenamiento'
    return 'Local'

def variant_label_en(variant_cn):
    if '客场' in variant_cn:
        return 'Away'
    if '特别版' in variant_cn:
        if 'Blanco' in variant_cn: return 'Special Ed. - White'
        if 'Azul' in variant_cn: return 'Special Ed. - Blue'
        if 'Rojo' in variant_cn: return 'Special Ed. - Red'
        if 'Negro' in variant_cn: return 'Special Ed. - Black'
        if 'Verde' in variant_cn: return 'Special Ed. - Green'
        if 'Rosa' in variant_cn: return 'Special Ed. - Pink'
        return 'Special Edition'
    if '训练服' in variant_cn or '赛前' in variant_cn:
        return 'Training'
    return 'Home'

def make_slug(team_es):
    """Convert Spanish team name to URL-safe slug"""
    slug = team_es.lower()
    replacements = {
        'á': 'a', 'é': 'e', 'í': 'i', 'ó': 'o', 'ú': 'u', 'ü': 'u',
        'ñ': 'n', ' ': '-', '_': '-'
    }
    for old, new in replacements.items():
        slug = slug.replace(old, new)
    slug = re.sub(r'[^a-z0-9-]', '', slug)
    slug = re.sub(r'-+', '-', slug).strip('-')
    return slug

# === MAIN ===
def load_excel_data(fname):
    """Read SKU data directly from Excel file"""
    import openpyxl
    excel_dir = os.path.join(BASE, '世界杯货盘')
    wb = openpyxl.load_workbook(os.path.join(excel_dir, fname))
    ws = wb['产品SKU表']
    data = []
    for row in range(2, ws.max_row + 1):
        sku_cn = ws.cell(row=row, column=2).value
        sku_en = ws.cell(row=row, column=3).value
        data.append({
            'row': row,
            'sku_cn': str(sku_cn) if sku_cn else '',
            'sku_en': str(sku_en) if sku_en else '',
        })
    wb.close()
    return data

def main():
    # Load SKU data directly from Excel files
    print('Reading Excel files...')
    sku2 = load_excel_data('球衣 (2).xlsx')
    sku5 = load_excel_data('球衣 (5).xlsx')
    print(f'  SKU data 2: {len(sku2)} rows')
    print(f'  SKU data 5: {len(sku5)} rows')

    # Build image file lists
    jersey2_files = {}
    if os.path.exists(JERSEY2_DIR):
        for f in sorted(os.listdir(JERSEY2_DIR)):
            if f.endswith('.jpg'):
                match = re.match(r'row(\d+)_', f)
                if match:
                    row = int(match.group(1))
                    jersey2_files[row] = os.path.join(JERSEY2_DIR, f)

    jersey5_files = {}
    if os.path.exists(JERSEY5_DIR):
        for f in sorted(os.listdir(JERSEY5_DIR)):
            if f.endswith('.jpg'):
                match = re.match(r'row(\d+)_', f)
                if match:
                    row = int(match.group(1))
                    jersey5_files[row] = os.path.join(JERSEY5_DIR, f)

    print(f'Jersey2 images: {len(jersey2_files)} rows')
    print(f'Jersey5 images: {len(jersey5_files)} rows')

    # Group by country
    # Structure: {country_cn: {variant_group: {variant_cn: [images]}}}
    countries = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))

    # Process 球衣 (2) data
    for item in sku2:
        sku_cn = item['sku_cn']
        if not sku_cn or sku_cn == 'None':
            continue
        country_cn, variant_cn = extract_country_and_variant(sku_cn)
        team_es = CN_TO_ES.get(country_cn.strip(), country_cn.strip())
        group = variant_to_group(variant_cn)

        # Get image
        row = item['row']
        if row in jersey2_files:
            countries[team_es][group][variant_cn].append({
                'row': row, 'source': 'jersey2', 'file': jersey2_files[row],
                'variant_cn': variant_cn
            })

    # Process 球衣 (5) data
    for item in sku5:
        sku_cn = item['sku_cn']
        if not sku_cn or sku_cn == 'None':
            continue
        country_cn, variant_cn = extract_country_and_variant(sku_cn)
        team_es = CN_TO_ES.get(country_cn.strip(), country_cn.strip())
        group = variant_to_group(variant_cn)

        row = item['row']
        if row in jersey5_files:
            # Find existing matching variant or add to same group
            found = False
            for vcn in countries[team_es][group]:
                if variant_cn in vcn or vcn in variant_cn:
                    countries[team_es][group][vcn].append({
                        'row': row, 'source': 'jersey5', 'file': jersey5_files[row],
                        'variant_cn': variant_cn
                    })
                    found = True
                    break
            if not found:
                countries[team_es][group][variant_cn].append({
                    'row': row, 'source': 'jersey5', 'file': jersey5_files[row],
                    'variant_cn': variant_cn
                })

    print(f'\nTotal countries found: {len(countries)}')

    # Build product entries
    national_products = []

    for team_es in sorted(countries.keys()):
        if not team_es or team_es == 'None':
            continue

        groups = countries[team_es]
        region = get_region(team_es)
        slug = make_slug(team_es)

        # Build default images (from home group, first variant)
        default_images = []

        # Build variants
        variants = []
        variant_order = ['home', 'away']

        for group_key in variant_order:
            if group_key not in groups or not groups[group_key]:
                continue

            group_variants = groups[group_key]

            # Merge all variant images within this group
            all_imgs = []
            for vcn, imgs in sorted(group_variants.items()):
                all_imgs.extend(imgs)

            # Sort: jersey2 images first, then jersey5
            imgs_2 = sorted([i for i in all_imgs if i['source'] == 'jersey2'], key=lambda x: x['row'])
            imgs_5 = sorted([i for i in all_imgs if i['source'] == 'jersey5'], key=lambda x: x['row'])

            # Merge: jersey2 first, jersey5 next
            merged = imgs_2 + imgs_5

            if not merged:
                continue

            # Copy images and build paths
            out_subdir = os.path.join(IMGS_DIR, f'jersey-{slug}')
            os.makedirs(out_subdir, exist_ok=True)

            img_paths = []
            for idx, img_info in enumerate(merged):
                src = img_info['file']
                is_model = img_info['source'] == 'jersey5'
                prefix = 'model' if is_model else 'main'
                fname = f'{group_key}_{prefix}_{idx+1:02d}.jpg'
                dst = os.path.join(out_subdir, fname)
                if not os.path.exists(dst):
                    shutil.copy2(src, dst)
                img_paths.append(f'images/mundial/jersey-{slug}/{fname}')

            if not default_images and img_paths:
                default_images = img_paths

            # Determine representative variant_cn for naming
            rep_vcn = sorted(group_variants.keys())[0] if group_variants else ('主场' if group_key == 'home' else '客场')
            labels_es = sorted(set(variant_label_es(i['variant_cn']) for i in merged if 'variant_cn' in i), key=len)
            label_es = labels_es[-1] if labels_es else ('Local' if group_key == 'home' else 'Visitante')
            labels_en = sorted(set(variant_label_en(i['variant_cn']) for i in merged if 'variant_cn' in i), key=len)
            label_en = labels_en[-1] if labels_en else ('Home' if group_key == 'home' else 'Away')

            variants.append({
                'id': group_key,
                'label': label_es,
                'label_en': label_en,
                'variant': rep_vcn,
                'images': img_paths
            })

        if not variants:
            continue

        if not default_images:
            default_images = variants[0]['images'] if variants else []

        # Build name
        name_es = f'Camiseta {team_es} 2026/27'
        name_en = f'{team_es} Jersey 2026/27'
        # Chinese team name (use first variant's first image source)
        team_zh = team_es  # fallback

        product = {
            'id': f'jersey-{slug}',
            'name_es': name_es,
            'name_en': name_en,
            'name_zh': f'{team_es} 2026/27',
            'category': 'jersey',
            'region': region,
            'team': team_es,
            'team_zh': team_zh,
            'variant': '主场',
            'season': '2026/27',
            'is_vintage': False,
            'sub_type': 'national',
            'sizes': ['S', 'M', 'L', 'XL', 'XXL'],
            'price': {
                'usd_small': 4.29,
                'usd_large': 3.71,
                'mxn_small': 75,
                'mxn_large': 65,
                'moq': 50
            },
            'images': default_images,
            'variants': variants,
            'raw_skus': [f'2627{team_es.lower()}{v["variant"]}' for v in variants],
            'source_files': ['球衣 (2).xlsx', '球衣 (5).xlsx'],
        }

        national_products.append(product)
        print(f'  {team_es}: {len(variants)} variants, {len(default_images)} default images')

    print(f'\nTotal national products: {len(national_products)}')

    # === Generate output ===
    # Read existing worldcup-products.js
    js_path = os.path.join(SITE_DIR, 'worldcup-products.js')
    with open(js_path, 'r', encoding='utf-8') as f:
        original_js = f.read()

    # Find all existing products
    # Parse all objects from the JS array
    all_objects = []
    depth = 0
    obj_start = -1
    for i, ch in enumerate(original_js):
        if ch == '{':
            if depth == 0:
                obj_start = i
            depth += 1
        elif ch == '}':
            depth -= 1
            if depth == 0 and obj_start >= 0:
                all_objects.append(original_js[obj_start:i+1])
                obj_start = -1

    print(f'\nParsed {len(all_objects)} existing products')

    # Categorize existing products
    non_national = []
    national_count = 0
    for obj_str in all_objects:
        cat_match = re.search(r'"category":\s*"([^"]*)"', obj_str)
        sub_match = re.search(r'"sub_type":\s*"([^"]*)"', obj_str)
        is_vintage_match = re.search(r'"is_vintage":\s*(true|false)', obj_str)

        is_national = (cat_match and cat_match.group(1) == 'jersey' and
                       sub_match and sub_match.group(1) == 'national' and
                       not (is_vintage_match and is_vintage_match.group(1) == 'true'))

        if is_national:
            national_count += 1
        else:
            non_national.append(obj_str)

    print(f'Existing national jerseys: {national_count}')
    print(f'Existing non-national products: {len(non_national)}')

    # Build new JS content
    new_national_json = ',\n '.join(json.dumps(p, ensure_ascii=False, indent=8) for p in national_products)

    # Rebuild the JS file
    new_js = 'window.WORLDCUP_PRODUCTS = [\n '
    new_js += new_national_json

    if non_national:
        new_js += ',\n '
        new_js += ',\n '.join(non_national)

    new_js += '\n];\n'

    # Write new JS file
    with open(js_path, 'w', encoding='utf-8') as f:
        f.write(new_js)

    print(f'\nWritten {len(national_products) + len(non_national)} total products to worldcup-products.js')
    print(f'  National: {len(national_products)}')
    print(f'  Club + Vintage + Non-jersey: {len(non_national)}')

    # === Save mapping for HTML modification ===
    mapping = {}
    for p in national_products:
        mapping[p['id']] = {
            'team': p['team'],
            'team_zh': p['team_zh'],
            'num_variants': len(p['variants']),
            'variant_labels': [v['label'] for v in p['variants']]
        }

    with open('/tmp/national_mapping.json', 'w', encoding='utf-8') as f:
        json.dump(mapping, f, ensure_ascii=False, indent=2)

    print('\nDone!')

if __name__ == '__main__':
    main()
